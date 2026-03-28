"""
Patent lookup — TJU Tech Transfer
Scrapes Google Patents directly. No API key, no credits, no signup.

HOW IT WORKS:
  Google Patents has predictable search URLs:
    https://patents.google.com/patent/US18700559  ← direct by serial (sometimes works)
    https://patents.google.com/results?q=...       ← search by title/inventor

  For each patent we:
    1. Try a direct URL if we can construct one (granted patents, some US apps)
    2. Search Google Patents by title + inventor keywords
    3. Scrape the result page for status, dates, assignee, link
    4. Provisionals (63/xxx, 62/xxx) are flagged — legally never published

PATENTSCOPE API (for PCT WO number lookup):
  Register free at: https://patentscope.wipo.int/patent/en/registration.jsf
  Then paste your username and password below.
  Leave blank to skip PCT WO number lookup.

REQUIREMENTS:
    pip install requests beautifulsoup4 pandas openpyxl lxml

USAGE:
    python patent_lookup.py
    (merged_patents.xlsx must be in the same folder)

OUTPUTS:
    patent_results.xlsx
    patent_summary.txt
"""

import re
import os
import sys
import time
import random
import subprocess
import getpass

# ── Dependency check — prompt to auto-install if anything is missing ──────────
_REQUIRED = {
    "requests":     "requests>=2.31.0",
    "pandas":       "pandas>=2.0.0",
    "bs4":          "beautifulsoup4>=4.12.0",
    "lxml":         "lxml>=5.0.0",
    "openpyxl":     "openpyxl>=3.1.0",
    "dotenv":       "python-dotenv>=1.0.0",
}
_missing = [pkg for pkg in _REQUIRED if __import__("importlib").util.find_spec(pkg) is None]
if _missing:
    print(f"\n⚠  Missing packages: {', '.join(_missing)}")
    ans = input("Install them now? [Y/n]: ").strip().lower()
    if ans in ("", "y", "yes"):
        subprocess.check_call([sys.executable, "-m", "pip", "install"] +
                              [_REQUIRED[p] for p in _missing])
        print("✓ Packages installed. Continuing...\n")
    else:
        print("Please install missing packages and re-run.\n"
              "  pip install -r requirements.txt\n"
              "  # or with conda:\n"
              "  conda install -c conda-forge requests beautifulsoup4 lxml pandas openpyxl python-dotenv\n"
              "  pip install playwright && playwright install chromium")
        sys.exit(1)

import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Playwright check (optional — only needed for design patents) ───────────────
_playwright_available = __import__("importlib").util.find_spec("playwright") is not None
if not _playwright_available:
    print("ℹ  Playwright not installed — design patent lookup (29/xxx) will use search link fallback.")
    print("   To enable: pip install playwright && playwright install chromium")
    print("   Or with conda: pip install playwright && playwright install chromium")
    print()

# Load credentials from .env — try multiple locations
try:
    from dotenv import load_dotenv
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    _cwd = os.getcwd()
    for _env_path in [
        os.path.join(_script_dir, ".env"),
        os.path.join(_cwd, ".env"),
        ".env",
    ]:
        if os.path.exists(_env_path):
            load_dotenv(_env_path, override=True)
            break
except ImportError:
    pass

INPUT_FILE  = "merged_patents.xlsx"
OUTPUT_XLSX = "patent_results.xlsx"
OUTPUT_TXT  = "patent_summary.txt"

# ── EPO OPS API CREDENTIALS ──────────────────────────────────────────────────
# Free registration at: https://developers.epo.org
# Create an app, get Consumer Key + Consumer Secret.
# Covers WO/PCT patents, US, EP and 100+ offices. 4GB/month free.
# Reads from .env first (EPO_KEY and EPO_SECRET), then prompts at runtime.
EPO_KEY    = os.getenv("EPO_KEY", "")
EPO_SECRET = os.getenv("EPO_SECRET", "")

if EPO_KEY:
    print(f"  ✓ EPO credentials loaded from .env")
else:
    EPO_KEY = input("EPO OPS Consumer Key (Enter to skip PCT lookup): ").strip()
if EPO_KEY and not EPO_SECRET:
    EPO_SECRET = getpass.getpass("EPO OPS Consumer Secret (hidden): ")

# ── ANTHROPIC API KEY (for NLP invention summaries) ───────────────────────────
# Get a free key at https://console.anthropic.com
# Set ANTHROPIC_API_KEY in your .env file, or leave blank to skip summaries.
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
if ANTHROPIC_API_KEY:
    print(f"  ✓ Anthropic API key loaded from .env")
else:
    ANTHROPIC_API_KEY = input("Anthropic API key for invention summaries (Enter to skip): ").strip()
    if ANTHROPIC_API_KEY:
        os.environ["ANTHROPIC_API_KEY"] = ANTHROPIC_API_KEY
# ─────────────────────────────────────────────────────────────────────────────

_epo_token = None
_epo_token_expiry = 0
# ─────────────────────────────────────────────────────────────────────────────

# Polite delay range in seconds — keeps Google from blocking us
DELAY_MIN = 3.0
DELAY_MAX = 6.0

# Rotate user agents so we look like a normal browser
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]

session = requests.Session()
scrape_count = 0


# ── Helpers ───────────────────────────────────────────────────────────────────

def clean(raw):
    if pd.isna(raw) or str(raw).strip() in ("", "nan", "NaN"):
        return None
    return str(raw).strip().replace("\u201c","").replace("\u201d","").replace('"',"").strip()


def is_provisional(serial):
    """63/xxx and 62/xxx are US provisionals — never published."""
    if not serial:
        return False
    return bool(re.match(r'^6[23]/', serial.strip()))


def is_pct(serial):
    return serial and serial.upper().startswith("PCT")


def sanitize_query(text):
    """Strip chars that break URL query strings."""
    text = re.sub(r'["\u201c\u201d\u2018\u2019]', '', text)
    text = re.sub(r'[/\\:;()\[\]{}<>!@#$%^*+=|~`]', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def keywords(title, n=7):
    """Extract n most meaningful words from title."""
    stop = {"and","the","for","with","from","that","this","into","via","using",
            "based","device","system","method","methods","apparatus","portable",
            "integrated","improved","novel"}
    words = [w for w in sanitize_query(title).lower().split()
             if len(w) > 3 and w not in stop]
    return " ".join(words[:n])


def get_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "DNT": "1",
        "Connection": "keep-alive",
    }


def polite_sleep():
    t = random.uniform(DELAY_MIN, DELAY_MAX)
    print(f"    (waiting {t:.1f}s...)")
    time.sleep(t)


def _filing_type(serial):
    """Human-readable filing type from serial number format."""
    if not serial:
        return "Unknown"
    s = serial.strip()
    if s.upper().startswith("PCT"):
        return "PCT (International)"
    if re.match(r'^6[23]/', s):
        return "Provisional (US)"
    if re.match(r'^29/', s):
        return "Design (US)"
    if re.match(r'^(1[5-9]|2\d)/', s):
        return "Utility (US)"
    if re.match(r'^(6\d|7\d|8\d|9\d|10\d|11\d|12\d|13\d|14\d)/', s):
        return "Utility (US)"
    return "US Application"


# ── Scraping ──────────────────────────────────────────────────────────────────

def fetch(url):
    """Fetch a URL, return BeautifulSoup or None."""
    global scrape_count
    try:
        r = session.get(url, headers=get_headers(), timeout=20)
        scrape_count += 1
        print(f"    GET {url[:80]} → {r.status_code}  [requests so far: {scrape_count}]")
        if r.status_code == 429:
            print("    Rate limited — sleeping 30s")
            time.sleep(30)
            return None
        if r.status_code != 200:
            return None
        return BeautifulSoup(r.text, "lxml")
    except Exception as e:
        print(f"    Fetch error: {e}")
        return None


def fetch_patent_page(patent_url):
    """
    Fetch a Google Patents patent page.
    Uses the XHR endpoint which returns pre-rendered HTML including
    all date fields (priority, filing, publication, expiration) that
    the regular page hides behind JavaScript.
    Falls back to regular URL if XHR fails.
    """
    # Convert standard URL to XHR endpoint
    # e.g. https://patents.google.com/patent/US11666728B2
    #   -> https://patents.google.com/xhr/result?id=patent/US11666728B2/en&exp=
    match = re.search(r'/patent/([^/]+)', patent_url)
    if match:
        pat_id = match.group(1)
        xhr_url = f"https://patents.google.com/xhr/result?id=patent/{pat_id}/en&exp="
        soup = fetch(xhr_url)
        if soup and soup.find("time"):   # XHR worked and has date elements
            return soup
    # Fallback to regular URL
    return fetch(patent_url)


def parse_patent_page(soup, url):
    """
    Extract fields from a Google Patents patent page.

    Google Patents renders content via Polymer web components — most fields
    are NOT in plain HTML. We extract data in this priority order:
      1. JSON-LD <script type="application/ld+json"> in <head>
      2. Embedded state JSON in <script> tags (key/value pairs)
      3. itemprop attributes as last fallback
    """
    if not soup:
        return None

    import json as _json

    # ── 1. JSON-LD block ───────────────────────────────────────────────────────
    ld = {}
    for script in soup.find_all("script", {"type": "application/ld+json"}):
        try:
            data = _json.loads(script.string or "")
            if isinstance(data, list):
                data = data[0]
            if isinstance(data, dict):
                ld = data
                break
        except Exception:
            continue

    # ── 2. Embedded patent state JSON ─────────────────────────────────────────
    # GP embeds a JS object like: var invention = {...} or window.invention = {...}
    # Also look for a script with recognisable patent fields
    state = {}
    for script in soup.find_all("script"):
        text = script.string or ""
        # Look for JSON blocks that contain patent-specific keys
        for pattern in (
            r'var\s+invention\s*=\s*(\{.*?\});',
            r'window\.invention\s*=\s*(\{.*?\});',
            r'"priorityDate"\s*:\s*"[^"]*"',  # presence check
        ):
            m = re.search(pattern, text, re.DOTALL)
            if m:
                # Try to extract the surrounding JSON object
                start = text.find('{', m.start())
                if start >= 0:
                    # Crude but reliable: find matching brace
                    depth, end = 0, start
                    for i, ch in enumerate(text[start:], start):
                        if ch == '{': depth += 1
                        elif ch == '}':
                            depth -= 1
                            if depth == 0:
                                end = i
                                break
                    try:
                        state = _json.loads(text[start:end+1])
                        break
                    except Exception:
                        pass
            if state:
                break

    def _get(*keys):
        """Try keys in ld then state, return first non-empty string."""
        for d in (ld, state):
            for k in keys:
                v = d.get(k, "")
                if v and str(v).strip() not in ("", "None", "null"):
                    return str(v).strip()
        return ""

    # ── 3. itemprop fallback helper ────────────────────────────────────────────
    def _itemprop(name):
        for tag in ("time", "dd", "span", "meta"):
            el = soup.find(tag, {"itemprop": name})
            if el:
                v = el.get("datetime") or el.get("content") or el.get_text(strip=True)
                if v:
                    return v
        return ""

    # ── Extract fields ─────────────────────────────────────────────────────────
    title = (
        _get("name", "title", "inventionTitle")
        or _itemprop("name")
        or (soup.find("h1", {"id": "title"}) or soup.find("span", {"itemprop": "title"}) or type("", (), {"get_text": lambda *a, **k: ""})()).get_text(strip=True)
    )

    status = _get("status", "legalStatus") or _itemprop("status")
    if not status:
        legal = soup.find("section", {"itemprop": "legalStatus"})
        if legal:
            status = legal.get_text(" ", strip=True)[:80]

    filing_date      = _get("filingDate", "filing_date")      or _itemprop("filingDate")
    priority_date    = _get("priorityDate", "priority_date")   or _itemprop("priorityDate")
    publication_date = _get("publicationDate", "grant_date", "publication_date") or _itemprop("publicationDate")

    # Expiration date — in the HTML as plain text near "Anticipated expiration" in the
    # legal status events section. Four fallback strategies, each more permissive.
    expiration_date = ""

    # Strategy 1: itemprop="anticipatedExpirationDate" on any tag
    expiration_date = _itemprop("anticipatedExpirationDate")

    # Strategy 2: find any element whose text contains "anticipated expiration"
    # and grab an adjacent <time> or date pattern
    if not expiration_date:
        for el in soup.find_all(True):
            txt = el.get_text(" ", strip=True)
            if "anticipated expiration" in txt.lower() and len(txt) < 200:
                t = el.find("time")
                if t:
                    expiration_date = t.get("datetime", t.get_text(strip=True))
                    break
                m = re.search(r'(\d{4}-\d{2}-\d{2})', txt)
                if m:
                    expiration_date = m.group(1)
                    break

    # Strategy 3: any <time> element whose nearby text mentions "expir"
    if not expiration_date:
        for t in soup.find_all("time"):
            nearby = (t.parent or t).get_text(" ", strip=True).lower()
            if "expir" in nearby or "anticipated" in nearby:
                expiration_date = t.get("datetime", t.get_text(strip=True))
                break

    # Strategy 4: scan script tags for the date near "anticipatedExpiration"
    if not expiration_date:
        for script in soup.find_all("script"):
            txt = script.string or ""
            m = re.search(r'[Aa]nticipated[Ee]xpiration[Dd]ate["\s:]+([0-9]{4}-[0-9]{2}-[0-9]{2})', txt)
            if m:
                expiration_date = m.group(1)
                break

    # Assignees
    assignees_orig    = [el.get_text(strip=True) for el in soup.find_all("dd", {"itemprop": "assigneeOriginal"})]
    assignees_current = [el.get_text(strip=True) for el in soup.find_all("dd", {"itemprop": "assigneeCurrent"})]
    # Also try JSON-LD
    if not assignees_orig:
        a = ld.get("assignee") or state.get("assignee") or ""
        if a:
            assignees_orig = [a] if isinstance(a, str) else [x.get("name", "") for x in a if isinstance(x, dict)]
    all_assignees = list(dict.fromkeys([a for a in assignees_orig + assignees_current if a]))
    assignee_str  = ", ".join(all_assignees)

    # Inventors
    inventors = [el.get_text(strip=True) for el in soup.find_all("dd", {"itemprop": "inventor"})]
    if not inventors:
        inv = ld.get("inventor") or state.get("inventor") or []
        if isinstance(inv, str):
            inventors = [inv]
        elif isinstance(inv, list):
            inventors = [x.get("name", x) if isinstance(x, dict) else str(x) for x in inv]
    inventor_str = ", ".join(inventors)

    # Abstract
    abstract = ""
    abs_el = soup.find("div", {"class": "abstract"}) or soup.find("section", {"itemprop": "abstract"})
    if abs_el:
        abstract = abs_el.get_text(" ", strip=True)[:500]
    if not abstract:
        abstract = str(_get("abstract", "description"))[:500]

    # Claims — first independent claim gives the clearest technical scope
    claims = ""
    claims_el = soup.find("section", {"itemprop": "claims"}) or soup.find("div", {"class": "claims"})
    if claims_el:
        claims = claims_el.get_text(" ", strip=True)[:1000]

    # Description — grab the first substantive paragraph of the detailed description
    description = ""
    desc_el = soup.find("section", {"itemprop": "description"}) or soup.find("div", {"class": "description"})
    if desc_el:
        # Skip boilerplate field-of-invention / background paragraphs, get first meaty paragraph
        paras = [p.get_text(" ", strip=True) for p in desc_el.find_all(["p","div"]) if len(p.get_text(strip=True)) > 100]
        description = " ".join(paras[:3])[:1000] if paras else desc_el.get_text(" ", strip=True)[:1000]

    def _fmt_date(raw):
        """Normalize any date string to YYYY-MM-DD, or return empty string."""
        if not raw:
            return ""
        s = str(raw).strip()
        # Strip trailing time component: "2022-10-11 00:00:00" -> "2022-10-11"
        s = s[:10]
        # Validate it looks like a date
        if re.match(r'^\d{4}-\d{2}-\d{2}$', s):
            return s
        return raw  # return as-is if we can't parse it

    filing_date      = _fmt_date(filing_date)
    priority_date    = _fmt_date(priority_date)
    publication_date = _fmt_date(publication_date)
    expiration_date  = _fmt_date(expiration_date)

    # WO publication number — only look in priority/family links, not citations
    # GP shows priority links like: "Priority to PCT/US2021/012888 priority patent/WO2021150386A1/en"
    wo_pub_num = ""
    # Look for priority section specifically
    priority_section = (
        soup.find("section", {"itemprop": "priority"}) or
        soup.find("ul", {"id": re.compile(r"priority", re.I)}) or
        soup.find("div", {"class": re.compile(r"priority", re.I)})
    )
    scan_area = priority_section if priority_section else None

    # Only scan priority links; if no priority section found, skip to avoid false positives
    if scan_area:
        for a in scan_area.find_all("a", href=True):
            m = re.search(r'/patent/(WO\d{7,}[A-Z]\d?)', a["href"], re.IGNORECASE)
            if m:
                wo_pub_num = m.group(1)
                break

    if not title and not filing_date and not priority_date:
        return None

    return {
        "gp_title":           title,
        "gp_status":          status,
        "gp_filing_date":     filing_date,
        "gp_priority_date":   priority_date,
        "gp_issue_date":      publication_date,
        "gp_expiration_date": expiration_date,
        "gp_assignee":        assignee_str,
        "gp_all_assignees":   all_assignees,
        "gp_inventor":        inventor_str,
        "gp_all_inventors":   inventors,
        "gp_abstract":        abstract,
        "gp_claims":          claims,
        "gp_description":     description,
        "gp_link":            url,
        "gp_wo_pub_num":      wo_pub_num,
    }


def search_google_patents(query):
    """
    Search Google Patents and return the first result URL, or None.
    Preserves slashes in PCT numbers since GP indexes them with slashes intact.
    """
    # Don't sanitize PCT numbers — they need their slashes
    if re.match(r'^PCT/', query.strip(), re.IGNORECASE):
        q = query.strip()
    else:
        q = sanitize_query(query)
    if not q:
        return None
    url = f"https://patents.google.com/?q={requests.utils.quote(q)}"
    soup = fetch(url)
    polite_sleep()
    if not soup:
        return None

    # Results are in search-result-item elements
    # Each has a data-result attribute with the patent ID
    results = soup.find_all("search-result-item") or soup.find_all(attrs={"data-result": True})
    if results:
        first = results[0]
        pat_id = first.get("data-result", "")
        if pat_id:
            return f"https://patents.google.com/patent/{pat_id}"

    # Fallback: find first patent link in page
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/patent/" in href and "google.com" not in href:
            return f"https://patents.google.com{href}"
        if "patents.google.com/patent/" in href:
            return href

    return None


def fetch_with_js(url, wait_for="networkidle", timeout=15000):
    """
    Fetch a URL using headless Playwright so JavaScript runs.
    Returns (final_url, BeautifulSoup) or (None, None) on failure.
    """
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto(url, wait_until=wait_for, timeout=timeout)
            final_url = page.url
            html = page.content()
            browser.close()
        return final_url, BeautifulSoup(html, "lxml")
    except ImportError:
        print("    Playwright not installed — run: pip install playwright && playwright install chromium")
        return None, None
    except Exception as e:
        print(f"    Playwright error: {e}")
        return None, None


def fetch_with_js_search(query, timeout=15000):
    """
    Use Playwright to type a query into Google Patents search box,
    wait for results, and return (final_url, soup).
    """
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.goto("https://patents.google.com", timeout=timeout)
            # Type into the search input and submit
            page.fill("input[name='q'], input[type='search'], #searchInput", query)
            page.keyboard.press("Enter")
            # Wait for navigation to complete
            page.wait_for_load_state("networkidle", timeout=timeout)
            final_url = page.url
            html = page.content()
            browser.close()
        return final_url, BeautifulSoup(html, "lxml")
    except ImportError:
        print("    Playwright not installed — run: pip install playwright && playwright install chromium")
        return None, None
    except Exception as e:
        print(f"    Playwright error: {e}")
        return None, None


def epo_get_abstract(wo_number):
    """
    Use EPO OPS API to get the English abstract for a WO publication number.
    Returns abstract string or "".
    """
    if not EPO_KEY or not EPO_SECRET:
        return ""
    token = epo_get_token()
    if not token:
        return ""
    try:
        # Normalize WO number to plain digits: WO 2022/082166 -> 2022082166
        wo_clean = re.sub(r'[^0-9A-Z]', '', wo_number.upper().replace("WO",""))
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}

        # Try docdb format first (WO.2022082166), then epodoc (WO2022082166)
        for fmt, doc_id in [("docdb", f"WO.{wo_clean}"), ("epodoc", f"WO{wo_clean}")]:
            url = f"https://ops.epo.org/3.2/rest-services/published-data/publication/{fmt}/{doc_id}/abstract"
            r = requests.get(url, headers=headers, timeout=15)
            print(f"    EPO abstract ({fmt}): {r.status_code}")
            if r.status_code != 200:
                continue
            data = r.json()
            # Navigate through exchange-documents wrapper
            doc = (data.get("ops:world-patent-data", {})
                       .get("exchange-documents", {})
                       .get("exchange-document", {}))
            if isinstance(doc, list):
                doc = doc[0]
            abstracts = doc.get("abstract", [])
            if isinstance(abstracts, dict):
                abstracts = [abstracts]
            for ab in abstracts:
                lang = ab.get("@lang", "")
                if lang.lower() == "en":
                    paras = ab.get("p", [])
                    if isinstance(paras, str):
                        return paras[:1500]
                    if isinstance(paras, dict):
                        return paras.get("$", "")[:1500]
                    if isinstance(paras, list):
                        return " ".join(p.get("$","") if isinstance(p,dict) else str(p) for p in paras)[:1500]
    except Exception as e:
        print(f"    EPO abstract error: {e}")
    return ""
    """
    Scrape abstract/claims from a Patentscope detail page using Playwright
    (since Patentscope is JS-rendered). Falls back to static fetch.
    """
    if not wipo_url or "patentscope" not in wipo_url:
        return None

    # Patentscope loads content via AJAX after a self-reload — must wait explicitly
    soup = None
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            # First load triggers a self-reload via setTimeout
            page.goto(wipo_url, wait_until="networkidle", timeout=20000)
            # Wait for the reload and AJAX content to settle
            page.wait_for_timeout(5000)
            # Wait for spinner to disappear
            try:
                page.wait_for_function(
                    "() => !document.querySelector('.b-view-panel__section')?.textContent?.includes('Loading')",
                    timeout=10000
                )
            except Exception:
                pass
            page.wait_for_timeout(2000)
            html = page.content()
            browser.close()
        soup = BeautifulSoup(html, "lxml")
    except Exception as e:
        print(f"    Playwright error: {e}")
    if not soup:
        soup = fetch(wipo_url)
    if not soup:
        return None

    def extract(s):
        abstract = ""
        claims = ""
        # Try known IDs/classes for abstract — prefer English version
        for sel in [{"id": "abstractEn"}, {"id": "abstractDiv"},
                    {"id": "abstract"}, {"class": "abstract"}, {"itemprop": "abstract"}]:
            el = s.find(True, sel)
            if el:
                txt = el.get_text(" ", strip=True)
                if len(txt) > 80:
                    abstract = txt[:1500]
                    break
        # Scan table cells for patent language — look for longest English cell
        if not abstract:
            candidates = []
            for td in s.find_all("td"):
                txt = td.get_text(" ", strip=True)
                if len(txt) > 150 and any(w in txt.lower() for w in [
                    "wherein", "comprising", "adapted", "configured",
                    "method", "system", "device", "apparatus", "disclosed"
                ]):
                    candidates.append(txt)
            if candidates:
                # Prefer the longest English candidate
                abstract = max(candidates, key=len)[:1500]
        # Scan divs/sections/paragraphs
        if not abstract:
            candidates = []
            for el in s.find_all(["div", "section", "p"]):
                txt = el.get_text(" ", strip=True)
                if len(txt) > 200 and any(w in txt.lower() for w in [
                    "wherein", "comprising", "adapted", "configured",
                    "method", "system", "device", "apparatus"
                ]):
                    candidates.append(txt)
            if candidates:
                abstract = max(candidates, key=len)[:1500]
        # Claims
        for sel in [{"id": "claimsEn"}, {"id": "claims"}, {"class": "claims"}]:
            el = s.find(True, sel)
            if el:
                txt = el.get_text(" ", strip=True)
                if len(txt) > 80:
                    claims = txt[:1500]
                    break
        return abstract, claims

    abstract, claims = extract(soup)
    print(f"    WIPO page length: {len(soup.get_text())} chars | abstract={len(abstract)} claims={len(claims)}")
    if abstract:
        print(f"    preview: {abstract[:100]}")

    if not any([abstract, claims]):
        return None

    return {"gp_abstract": abstract, "gp_claims": claims, "gp_description": ""}



def generate_invention_summary(family_title, patent_number, abstract, claims, description):
    """
    Use Claude API to generate a concise invention summary from patent text.
    Returns a string with technical summary followed by plain English summary.
    Returns "" if no text available or API call fails.
    """
    # Design patents have no technical description — just figure captions and a single claim
    # Detect this case and return a standard note
    is_design = (
        not abstract and
        claims and "ornamental design" in claims.lower() and
        len(claims) < 200
    )
    if is_design:
        m = re.search(r'ornamental design for (.+?)(?:,|\.|$)', claims, re.IGNORECASE)
        subject = m.group(1).strip() if m else family_title
        return (f"Technical: US design patent protecting the ornamental appearance of {subject}. "
                f"Design patents cover visual aesthetics, not functional features.\n"
                f"Plain: This is a design patent protecting how the {subject} looks, not how it works.")

    # Collect available text
    text_parts = []
    if abstract:    text_parts.append(f"ABSTRACT:\n{abstract}")
    if claims:      text_parts.append(f"CLAIMS:\n{claims}")
    if description: text_parts.append(f"DESCRIPTION:\n{description}")
    if not text_parts:
        return ""

    # Use first available as primary if abstract missing
    patent_text = "\n\n".join(text_parts)
    context = f"Based on the following patent text:\n\n{patent_text}"
    prompt = f"""You are summarizing a patent for a technology transfer office.

Patent: {family_title}
Patent Number: {patent_number or 'Pending'}

{context}

Write a two-part summary:

1. TECHNICAL: One concise sentence describing the core technical innovation for an IP/legal audience.
2. PLAIN: One concise sentence describing the invention in plain English for a business/licensing audience.

Format exactly as:
Technical: [sentence]
Plain: [sentence]

Be specific to this invention. Do not use generic phrases like "novel approach" or "innovative method".
If the provided text does not contain enough technical information to write a meaningful summary, respond with exactly:
Technical: N/A — insufficient patent text available.
Plain: N/A — insufficient patent text available."""

    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": os.getenv("ANTHROPIC_API_KEY",""), "anthropic-version": "2023-06-01", "content-type": "application/json"},
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 300, "messages": [{"role": "user", "content": prompt}]},
            timeout=30
        )
        if r.status_code == 200:
            return r.json()["content"][0]["text"].strip()
        else:
            print(f"    Claude API: {r.status_code}")
    except Exception as e:
        print(f"    Claude API error: {e}")
    return ""


def normalize_pct(serial):
    """Normalize PCT serial to 4-digit year + zero-padded sequence."""
    m = re.match(r'PCT/([A-Z]{2})(\d{2,4})/(\d+)', serial.strip(), re.IGNORECASE)
    if not m:
        return serial.strip()
    cc, yr, seq = m.group(1).upper(), m.group(2), m.group(3)
    if len(yr) == 2:
        yr = '20' + yr
    return f"PCT/{cc}{yr}/{seq.zfill(6)}"


def uspto_get_design_number(serial_29):
    """
    Use the free USPTO ODP API to get the granted patent number for a
    design application (29/xxx). No API key required.
    Returns e.g. "D874011" or "".
    """
    try:
        # Format: 29/666,847 -> 29666847
        n = re.sub(r'[^0-9]', '', serial_29)
        url = f"https://api.patentsview.org/patents/query"
        params = {
            "q": f'{{"_eq":{{"app_number":"{n}"}}}}',
            "f": '["patent_number","patent_type"]',
            "o": '{"per_page":1}'
        }
        r = requests.get(url, params=params, timeout=15)
        print(f"    USPTO ODP: {r.status_code}")
        if r.status_code == 200:
            data = r.json()
            patents = data.get("patents") or []
            if patents:
                pat_num = patents[0].get("patent_number", "")
                pat_type = patents[0].get("patent_type", "")
                print(f"    USPTO found: {pat_num} ({pat_type})")
                if pat_num:
                    return f"D{pat_num}" if not pat_num.startswith("D") else pat_num
    except Exception as e:
        print(f"    USPTO ODP error: {e}")
    return ""
    """
    Normalize a PCT serial to full 4-digit year and zero-padded sequence.
    PCT/US23/75015  -> PCT/US2023/075015
    """
    m = re.match(r'PCT/([A-Z]{2})(\d{2,4})/(\d+)', serial.strip(), re.IGNORECASE)
    if not m:
        return serial.strip()
    cc, yr, seq = m.group(1).upper(), m.group(2), m.group(3)
    if len(yr) == 2:
        yr = '20' + yr
    seq = seq.zfill(6)
    return f"PCT/{cc}{yr}/{seq}"


def epo_get_token():
    """Get/refresh EPO OPS OAuth2 token. Cached until expiry."""
    global _epo_token, _epo_token_expiry
    if _epo_token and time.time() < _epo_token_expiry - 60:
        return _epo_token
    try:
        r = requests.post(
            "https://ops.epo.org/3.2/auth/accesstoken",
            data={"grant_type": "client_credentials"},
            auth=(EPO_KEY, EPO_SECRET),
            timeout=10
        )
        if r.status_code == 200:
            data = r.json()
            _epo_token = data["access_token"]
            _epo_token_expiry = time.time() + int(data.get("expires_in", 1200))
            return _epo_token
        print(f"    EPO token error: {r.status_code}")
    except Exception as e:
        print(f"    EPO token error: {e}")
    return None


def epo_get_wo(pct_normalized):
    """
    Use EPO OPS API to find the WO publication number for a PCT application.
    Calls the family endpoint with the PCT application number to get related
    WO publications.
    Returns "WO YYYY/NNNNN" or "".
    """
    if not EPO_KEY or not EPO_SECRET:
        return ""
    token = epo_get_token()
    if not token:
        return ""
    try:
        # Convert PCT/US2023/075015 to EPO docdb format: US.2023075015.A
        m = re.match(r'PCT/([A-Z]{2})(\d{4})/(\d+)', pct_normalized, re.IGNORECASE)
        if not m:
            return ""
        cc, yr, seq = m.group(1).upper(), m.group(2), m.group(3)
        # EPO application number format for PCT: CC.YYYYseq (no slash, no PCT/)
        epo_appnum = f"{cc}.{yr}{seq}"

        headers = {
            "Authorization": f"Bearer {token}",
            "Accept": "application/json",
        }
        # Search family by application number
        url = f"https://ops.epo.org/3.2/rest-services/family/application/docdb/{epo_appnum}"
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code == 404:
            return ""
        if r.status_code != 200:
            print(f"    EPO OPS: {r.status_code}")
            return ""

        data = r.json()
        # Walk patent family members looking for WO publications
        families = data.get("ops:world-patent-data", {}) \
                       .get("ops:patent-family", {}) \
                       .get("ops:family-member", [])
        if isinstance(families, dict):
            families = [families]

        for member in families:
            pubs = member.get("publication-reference", [])
            if isinstance(pubs, dict):
                pubs = [pubs]
            for pub in pubs:
                doc = pub.get("document-id", {})
                if isinstance(doc, list):
                    doc = doc[0]
                country = doc.get("country", {}).get("$", "")
                if country.upper() == "WO":
                    doc_num = doc.get("doc-number", {}).get("$", "")
                    year = doc.get("date", {}).get("$", "")[:4]
                    if doc_num:
                        # Format as WO YYYY/NNNNN
                        m2 = re.match(r'(\d{4})(\d+)', doc_num)
                        if m2:
                            return f"WO {m2.group(1)}/{m2.group(2)}"
                        return f"WO {year}/{doc_num}" if year else f"WO {doc_num}"
    except Exception as e:
        print(f"    EPO OPS error: {e}")
    return ""


def epo_get_wo_title(wo_number):
    """
    Fetch the English title for a WO publication from EPO OPS.
    Returns title string or "".
    """
    if not EPO_KEY or not EPO_SECRET or not wo_number:
        return ""
    token = epo_get_token()
    if not token:
        return ""
    try:
        wo_clean = re.sub(r'[^0-9A-Z]', '', wo_number.upper().replace("WO",""))
        url = f"https://ops.epo.org/3.2/rest-services/published-data/publication/docdb/WO.{wo_clean}/biblio"
        headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
        r = requests.get(url, headers=headers, timeout=15)
        if r.status_code != 200:
            return ""
        data = r.json()
        doc = (data.get("ops:world-patent-data", {})
                   .get("exchange-documents", {})
                   .get("exchange-document", {}))
        if isinstance(doc, list):
            doc = doc[0]
        titles = doc.get("bibliographic-data", {}).get("invention-title", [])
        if isinstance(titles, dict):
            titles = [titles]
        for t in titles:
            if t.get("@lang", "").lower() == "en":
                return t.get("$", "")
    except Exception:
        pass
    return ""


def direct_gp_url(serial, patent_num, filing_type=""):
    """
    Try to build a direct Google Patents URL.
    For design patents where patent_num == serial (bad data), returns None
    so the script falls back to EPO lookup or title search.
    """
    # Ignore patent_num if it's the same as the serial — data entry error
    if patent_num and patent_num.strip() == (serial or "").strip():
        patent_num = None

    # Design patent with a real granted D-number (e.g. "D874,011")
    if patent_num and re.match(r'^D\d', patent_num.strip(), re.IGNORECASE):
        n = re.sub(r'[^0-9]', '', patent_num)
        return f"https://patents.google.com/patent/USD{n}S1"

    # Design patent serial (29/xxx) used as patent number — skip, handled separately
    if filing_type == "Design (US)" and (not patent_num or patent_num.startswith("29/")):
        return None

    if patent_num:
        n = patent_num.replace(",","").replace(" ","")
        return f"https://patents.google.com/patent/US{n}B2"

    if serial and not is_pct(serial) and not is_provisional(serial):
        n = re.sub(r"[,/ ]", "", serial)
        return f"https://patents.google.com/patent/US{n}"
    return None


def uspto_get_design_pub_number(serial_29):
    """
    Use the free USPTO ODP API to find the granted patent number for a
    design application (29/xxx serial). Returns e.g. "D874011" or "".
    No API key required.
    """
    try:
        # Normalize: 29/666,847 -> 29666847
        n = re.sub(r'[^0-9]', '', serial_29)
        # USPTO ODP application endpoint
        url = f"https://api.patentsview.org/patents/query"
        params = {
            "q": f'{{"_eq":{{"app_number":"{n}"}}}}',
            "f": '["patent_number","patent_type","app_number"]',
            "o": '{"per_page":1}'
        }
        r = requests.get(url, params=params, timeout=15)
        if r.status_code == 200:
            data = r.json()
            patents = data.get("patents") or []
            if patents:
                pat_num = patents[0].get("patent_number","")
                pat_type = patents[0].get("patent_type","")
                if pat_num and "design" in pat_type.lower():
                    return f"D{pat_num}"
    except Exception as e:
        print(f"    USPTO ODP error: {e}")

    # Fallback: try USPTO PEDS API
    try:
        n_fmt = serial_29.strip()  # keep original format e.g. 29/666,847
        url2 = f"https://ped.uspto.gov/api/queries"
        payload = {
            "searchText": f"applId:({re.sub(r'[^0-9]', '', n_fmt)})",
            "fl": "applId,patentNumber,appType",
            "rows": 1
        }
        r2 = requests.post(url2, json=payload, timeout=15)
        if r2.status_code == 200:
            data2 = r2.json()
            docs = data2.get("queryResults",{}).get("searchResponse",{}).get("response",{}).get("docs",[])
            if docs:
                pat_num = docs[0].get("patentNumber","")
                if pat_num:
                    return f"D{pat_num}" if not pat_num.startswith("D") else pat_num
    except Exception:
        pass
    return ""
    return None


# ── Main lookup ───────────────────────────────────────────────────────────────

def lookup(row):
    serial     = clean(row.get("Serial Number"))
    patent_num = clean(row.get("Patent Number"))
    pub_num    = clean(row.get("Publication Number"))
    title      = str(row.get("Title","")).strip()
    inventors  = str(row.get("Inventors","")).strip()
    inv_parts  = inventors.split(",")[0].strip().split() if inventors not in ("","nan") else []
    last_name  = inv_parts[-1] if inv_parts else ""
    kw         = keywords(title)

    # ── Provisionals: legally never published ─────────────────────────────────
    if is_provisional(serial):
        print(f"  ⚠ Provisional — not publicly available by law")
        return {
            "gp_title":"", "gp_status":"Provisional — not published",
            "gp_filing":"", "gp_grant":"", "gp_assignee":"",
            "gp_all_assignees":[], "gp_inventor":"", "gp_all_inventors":[],
            "gp_abstract":"", "gp_link":"",
            "not_found_reason": "Provisional applications (63/xxx, 62/xxx) are never published by US patent law.",
        }, "provisional_unpublished"

    # ── PCT: search Google Patents + extract WO publication number ─────────────
    if is_pct(serial):
        wipo_link = ""
        wo_pub_num = ""
        # Normalize to 4-digit year + zero-padded sequence for all lookups
        serial_norm = normalize_pct(serial)

        # Build WIPO link if we already have the pub number
        if pub_num and pub_num.upper().startswith("WO"):
            wo_pub_num = pub_num
            wipo_id = re.sub(r"[/ ]", "", pub_num.upper())
            wipo_link = f"https://patentscope.wipo.int/search/en/detail.jsf?docId={wipo_id}"

        # Try GP search with normalized PCT number, then title+inventor
        search_queries = [serial_norm]
        if pub_num and pub_num.upper().startswith("WO"):
            search_queries.append(sanitize_query(pub_num))
        search_queries.append(f"{kw} {last_name}".strip() if last_name else kw)

        gp_url = None
        for q in search_queries:
            if not q:
                continue
            print(f"  PCT — searching GP: {q[:60]}")
            gp_url = search_google_patents(q)
            if gp_url:
                break

        if gp_url:
            print(f"  Fetching GP page: {gp_url}")
            soup = fetch_patent_page(gp_url)
            polite_sleep()
            result = parse_patent_page(soup, gp_url)
            if result:
                # Extract WO publication number from the GP URL if we didn't have it
                if not wo_pub_num:
                    wo_match = re.search(r'/patent/(WO\d+[A-Z0-9]*)', gp_url, re.IGNORECASE)
                    if wo_match:
                        wo_pub_num = wo_match.group(1)
                        # Format nicely: WO2024012345A1 -> WO 2024/012345
                        m2 = re.match(r'WO(\d{4})(\d+)([A-Z]\d+)?', wo_pub_num, re.IGNORECASE)
                        if m2:
                            wo_pub_num = f"WO {m2.group(1)}/{m2.group(2)}"

                # Build WIPO link from found WO number if we didn't have one
                if wo_pub_num and not wipo_link:
                    wipo_id = re.sub(r"[/ ]", "", wo_pub_num.upper())
                    wipo_link = f"https://patentscope.wipo.int/search/en/detail.jsf?docId={wipo_id}"

                if wipo_link and not result["gp_link"]:
                    result["gp_link"] = wipo_link
                result["not_found_reason"] = ""
                result["gp_wo_pub_num"] = wo_pub_num
                print(f"  ✓ PCT found on GP: '{result['gp_title'][:50]}' | WO: {wo_pub_num}")
                return result, "gp_search_pct"

        if wipo_link:
            return {
                "gp_title":"", "gp_status": str(row.get("Status","")),
                "gp_filing_date": str(row.get("File Date","")),
                "gp_priority_date":"", "gp_issue_date":"", "gp_expiration_date":"",
                "gp_assignee":"", "gp_all_assignees":[],
                "gp_inventor":"", "gp_all_inventors":[], "gp_abstract":"",
                "gp_link": wipo_link,
                "gp_wo_pub_num": wo_pub_num,
                "not_found_reason": "PCT found on WIPO only — no Google Patents page scraped.",
            }, "wipo_link_only"

        # ── Patentscope API + Espacenet scrape — get WO number ────────────────
        if serial_norm and not wo_pub_num:
            wo_title = ""  # title from WO publication, for mismatch detection
            # Try EPO OPS API first (requires free key from developers.epo.org)
            if EPO_KEY and EPO_SECRET:
                print(f"  Trying EPO OPS API: {serial_norm}")
                wo_pub_num = epo_get_wo(serial_norm)
                if wo_pub_num:
                    print(f"  ✓ EPO OPS found: {wo_pub_num}")
                    # Also fetch the WO title so we can detect mismatches later
                    wo_title = epo_get_wo_title(wo_pub_num)

            # Try Patentscope web search with normalized number (no key needed)
            if not wo_pub_num:
                ps_url = f"https://patentscope.wipo.int/search/en/search.jsf?query=AN%3A{requests.utils.quote(serial_norm)}&office=WO"
                print(f"  Trying Patentscope web: {serial_norm}")
                ps_soup = fetch(ps_url)
                polite_sleep()
                if ps_soup:
                    page_text = ps_soup.get_text(" ")
                    m = re.search(r'WO[/ ]?(\d{4})[/ ](\d{5,})', page_text)
                    if m:
                        wo_pub_num = f"WO {m.group(1)}/{m.group(2)}"

            # Try Espacenet with converted format
            if not wo_pub_num:
                m = re.match(r'PCT/([A-Z]{2})(\d{4})/(\d+)', serial_norm, re.IGNORECASE)
                if m:
                    cc, yr, seq = m.group(1).upper(), m.group(2), m.group(3).lstrip('0') or '0'
                    espacenet_num = f"WO{yr}{cc}{seq}"
                    espacenet_url = f"https://worldwide.espacenet.com/patent/search/family/000000000/publication/{espacenet_num}"
                    print(f"  Trying Espacenet: {espacenet_num}")
                    esp_soup = fetch(espacenet_url)
                    polite_sleep()
                    if esp_soup:
                        page_text = esp_soup.get_text(" ")
                        m2 = re.search(r'WO[/ ]?(\d{4})[/ ](\d{5,})', page_text)
                        if m2:
                            wo_pub_num = f"WO {m2.group(1)}/{m2.group(2)}"
                        if not wo_pub_num:
                            for a in esp_soup.find_all("a", href=True):
                                m3 = re.search(r'(WO\d{7,}[A-Z]\d?)', a["href"] + a.get_text(), re.IGNORECASE)
                                if m3:
                                    raw = m3.group(1).upper()
                                    m4 = re.match(r'WO(\d{4})(\d+)', raw)
                                    if m4:
                                        wo_pub_num = f"WO {m4.group(1)}/{m4.group(2)}"
                                        break

            if wo_pub_num:
                wipo_id = re.sub(r"[/ ]", "", wo_pub_num.upper())
                wipo_link = f"https://patentscope.wipo.int/search/en/detail.jsf?docId={wipo_id}"
                print(f"  ✓ Found WO number: {wo_pub_num}")
                return {
                    "gp_title": wo_title if wo_title else "", 
                    "gp_status": str(row.get("Status","")),
                    "gp_filing_date": str(row.get("File Date","")),
                    "gp_priority_date":"", "gp_issue_date":"", "gp_expiration_date":"",
                    "gp_assignee":"", "gp_all_assignees":[],
                    "gp_inventor":"", "gp_all_inventors":[], "gp_abstract":"",
                    "gp_link": wipo_link,
                    "gp_wo_pub_num": wo_pub_num,
                    "not_found_reason": "",
                }, "patentscope_scrape"

        return {
            "gp_wo_pub_num": "",
            "not_found_reason": "PCT application — no WIPO publication number available and title search returned no results.",
        }, "not_found"

    # ── Try direct URL first ───────────────────────────────────────────────────
    filing_type_str = _filing_type(serial)

    # For design patents with bad/missing patent number, use EPO to get granted D number
    # For design patents, just build the GP search link directly — no fetching
    if filing_type_str == "Design (US)" and serial and serial.startswith("29/"):
        if not patent_num or patent_num.strip() == serial.strip():
            n = re.sub(r'[^0-9]', '', serial.split('/', 1)[-1])
            search_query = f"29/{n}"
            search_url = f"https://patents.google.com/?q={requests.utils.quote(search_query)}"
            print(f"  Design patent — typing into GP search: {search_query}")
            final_url, soup = fetch_with_js_search(search_query)
            polite_sleep()
            if final_url and "/patent/" in final_url:
                # GP redirected directly to the patent page
                print(f"  JS redirected to: {final_url}")
                soup2 = fetch_patent_page(final_url)
                polite_sleep()
                result = parse_patent_page(soup2, final_url)
                if result:
                    result["not_found_reason"] = ""
                    print(f"  ✓ Design patent found: '{result['gp_title'][:50]}'")
                    return result, "design_direct"
                # After JS runs, look for patent URL in rendered HTML
                for a in soup.find_all("a", href=True):
                    href = a["href"]
                    if "/patent/USD" in href or re.search(r'/patent/US\w+S\d', href):
                        design_url = href if href.startswith("http") else f"https://patents.google.com{href}"
                        print(f"  Found design patent URL: {design_url}")
                        soup2 = fetch_patent_page(design_url)
                        polite_sleep()
                        result = parse_patent_page(soup2, design_url)
                        if result:
                            result["not_found_reason"] = ""
                            print(f"  ✓ Design patent found: '{result['gp_title'][:50]}'")
                            return result, "design_direct"
                # Check if page itself is the patent page
                result = parse_patent_page(soup, final_url or search_url)
                if result and result.get("gp_title"):
                    result["not_found_reason"] = ""
                    print(f"  ✓ Design patent found (page): '{result['gp_title'][:50]}'")
                    return result, "design_direct"
        n = re.sub(r'[^0-9]', '', serial.split('/', 1)[-1])
        gp_link = f"https://patents.google.com/?q=29/{n}"
        print(f"  Design patent — using search link fallback")
        return {
            "gp_title": title, "gp_status": str(row.get("Status", "")),
            "gp_filing_date": "", "gp_priority_date": "",
            "gp_issue_date": "", "gp_expiration_date": "",
            "gp_assignee": "", "gp_all_assignees": [],
            "gp_inventor": "", "gp_all_inventors": [],
            "gp_abstract": "", "gp_link": gp_link,
            "gp_wo_pub_num": "", "not_found_reason": "Design patent — search link only, click to find granted patent",
        }, "design_search"

    direct = direct_gp_url(serial, patent_num, filing_type_str)
    if direct:
        print(f"  Trying direct URL: {direct}")
        soup = fetch_patent_page(direct)
        polite_sleep()
        result = parse_patent_page(soup, direct)
        if result:
            result["not_found_reason"] = ""
            print(f"  ✓ Direct URL worked: '{result['gp_title'][:50]}' | status: '{result['gp_status']}'")
            return result, "direct_url"
        print(f"  Direct URL gave no data — falling back to search")

    # ── Search by title + inventor ─────────────────────────────────────────────
    query = f"{kw} {last_name}".strip() if last_name else kw
    print(f"  Searching GP: {query[:70]}")
    gp_url = search_google_patents(query)
    if gp_url:
        print(f"  Fetching GP page: {gp_url}")
        soup = fetch_patent_page(gp_url)
        polite_sleep()
        result = parse_patent_page(soup, gp_url)
        if result:
            result["not_found_reason"] = ""
            print(f"  ✓ Search found: '{result['gp_title'][:50]}' | status: '{result['gp_status']}'")
            return result, "gp_title_search"

    reason = (
        "Application may not yet be published (US apps publish 18 months after filing). "
        "Direct URL returned 404 and title/inventor search returned no matching results."
        if serial else
        "No serial number in source data and no title match found on Google Patents."
    )
    print(f"  ✗ Not found")
    return {"not_found_reason": reason}, "not_found"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    import json

    CACHE_FILE = "patent_cache.json"

    print(f"Reading {INPUT_FILE}...")

    # ── Auto-detect file format ────────────────────────────────────────────────
    # Format A (merged_patents.xlsx): has proper headers including "Tech ID",
    #   "Serial Number", "Title", "Inventors", etc. One row per patent application.
    # Format B (selected-tec.xlsx): no header row, one row per invention/technology.
    #   Columns: 0=Tech ID, 1=Title, 2=Licensing Manager, 3=Secondary Manager,
    #            4=Inventor, 5=Status, 6=date, 7=date, 8=Patented(Y/N), ...

    raw = pd.read_excel(INPUT_FILE, dtype=str, header=None)
    first_cell = str(raw.iloc[0, 0]).strip()

    # If first row looks like a Tech ID (e.g. "LEO_JOS.001", "JeffSolves22.001")
    # rather than a column label, it's Format B
    is_format_b = bool(re.match(r'^[A-Z_]+\d|^Jeff', first_cell, re.IGNORECASE)) \
                  or "Tech ID" not in raw.iloc[0].tolist()

    if is_format_b:
        print("  Detected format: Technology list (no header row, one row per invention)")
        df = raw.copy()
        # Map known column positions to standard names
        col_map = {
            0: "Tech ID",
            1: "Title",
            2: "Licensing Manager First Name",   # stored as "Last, First" but we'll use as-is
            4: "Inventors",
            5: "Status (Internal)",
            8: "Patented",
        }
        df = df.rename(columns=col_map)
        df["Serial Number"] = ""   # no serial numbers in this format
        df["Patent Number"] = ""
        df["Publication Number"] = ""
        df["File Date"] = ""
        df["Status Date (Internal)"] = ""
        df["Licensing Manager Last Name"] = ""
        df["Country"] = "United States"
        # Clean up Licensing Manager — it's "Last, First" format, flip it
        df["Licensing Manager First Name"] = df["Licensing Manager First Name"].apply(
            lambda v: " ".join(reversed(str(v).split(", "))) if ", " in str(v) else str(v)
        )
        format_label = "B (technology list)"
    else:
        print("  Detected format: Patent applications list (with header row)")
        df = pd.read_excel(INPUT_FILE, dtype=str)
        format_label = "A (patent applications)"

    print(f"  {len(df)} rows loaded  [format {format_label}]")

    # Dedup — for format B there are no duplicates expected, but run anyway
    dedup_cols = ["Tech ID", "Title"]
    if "Serial Number" in df.columns:
        dedup_cols.append("Serial Number")
    df_u = df.drop_duplicates(subset=dedup_cols).reset_index(drop=True)
    n_dupes = len(df) - len(df_u)
    print(f"  {n_dupes} duplicate rows removed → {len(df_u)} unique records.\n")

    # Load cache — keyed by serial number, avoids re-scraping on reruns
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            print(f"  Loaded {len(cache)} cached records from {CACHE_FILE}\n")
        except Exception:
            cache = {}

    # Helper: strip time component from any date string coming out of pandas
    def _d(raw):
        s = str(raw or "").strip()
        if not s or s == "nan":
            return ""
        # Strip " 00:00:00" or "T00:00:00" suffixes
        s = re.sub(r'[\sT]\d{2}:\d{2}:\d{2}.*$', '', s).strip()
        return s

    rows_out = []
    for i, row in df_u.iterrows():
        serial    = clean(row.get("Serial Number"))
        title     = str(row.get("Title","")).strip()
        inv       = str(row.get("Inventors","")).strip()
        first_inv = inv.split(",")[0].strip() if inv not in ("","nan") else ""
        cache_key = serial or title[:60]

        print(f"[{i+1}/{len(df_u)}] {row.get('Tech ID','')} | {serial or 'NO SERIAL'}")
        print(f"  Title: {title[:65]}")

        if cache_key in cache:
            print(f"  ✓ Using cached result")
            gp     = cache[cache_key]["gp"]
            method = cache[cache_key]["method"]
        else:
            gp, method = lookup(row)
            cache[cache_key] = {"gp": gp, "method": method}
            # Save cache after each new scrape so crashes don't lose work
            try:
                with open(CACHE_FILE, "w", encoding="utf-8") as f:
                    json.dump(cache, f, indent=2, ensure_ascii=False)
            except Exception:
                pass

        rows_out.append({
            "Tech ID":               row.get("Tech ID",""),
            "Family Size":           "",   # filled after loop
            "Filing Type":           _filing_type(serial),
            "Title (Original)":      title,
            "Serial Number":         serial or "",
            "Country":               row.get("Country",""),
            "Status (Internal)":     row.get("Status",""),
            "Status Date (Internal)": _d(row.get("Status Date","")),
            "Patent Number":         str(row.get("Patent Number","") or "").strip().replace("nan",""),
            "Publication Number":    str(row.get("Publication Number","") or "").strip().replace("nan",""),
            "File Date":             _d(row.get("File Date","")),
            "Inventors":             first_inv,
            "Licensing Manager":     f"{row.get('Licensing Manager First Name','')} {row.get('Licensing Manager Last Name','')}".strip(),
            "GP Title":              gp.get("gp_title",""),
            "GP Status":             gp.get("gp_status",""),
            "GP Filing Date":        gp.get("gp_filing_date",""),
            "GP Priority Date":      gp.get("gp_priority_date",""),
            "GP Issue Date":         gp.get("gp_issue_date",""),
            "GP Est. Exp. Date":          gp.get("gp_expiration_date",""),
            "GP Assignee":           gp.get("gp_assignee",""),
            "GP All Assignees":      ", ".join(gp.get("gp_all_assignees",[])),
            "GP Inventors":          gp.get("gp_inventor",""),
            "GP All Inventors":      ", ".join(gp.get("gp_all_inventors",[])),
            "GP WO Pub. No.":        gp.get("gp_wo_pub_num",""),
            "Google Patents Link":   gp.get("gp_link",""),
            "Search Method":         method,
            "Not Found Reason":      gp.get("not_found_reason",""),
            "Abstract (snippet)":    gp.get("gp_abstract",""),
            "gp_claims":             gp.get("gp_claims",""),
            "gp_description":        gp.get("gp_description",""),
        })
        print()

    # Save cache after every run
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(cache, f, indent=2, ensure_ascii=False)
        print(f"Cache saved → {CACHE_FILE}  ({len(cache)} entries)")
    except Exception as e:
        print(f"  Cache save failed: {e}")

    # ── Back-fill Family Size now that all rows are collected ──────────────────
    family_sizes = {}
    for r in rows_out:
        tid = r["Tech ID"]
        family_sizes[tid] = family_sizes.get(tid, 0) + 1
    for r in rows_out:
        r["Family Size"] = family_sizes[r["Tech ID"]]

    # Sort by Tech ID so family members are grouped together
    rows_out.sort(key=lambda r: (r["Tech ID"], r["File Date"] or ""))

    # ── Mismatch detection — clear GP data for rows where wrong patent was found ─
    def _inventors_overlap(inv1, inv2):
        if not inv1 or not inv2:
            return False
        def lastnames(s):
            parts = re.split(r'[;,]', str(s).lower())
            return set(re.sub(r'[^a-z]','',p.strip().split()[-1]) for p in parts if p.strip())
        return bool(lastnames(inv1) & lastnames(inv2))

    def _is_mismatch(r):
        internal = r.get("Title (Original)","")
        gp       = r.get("GP Title","")
        if not gp or not r.get("Google Patents Link",""):
            return False
        stops = {'a','an','the','and','or','of','for','to','in','with','on','by','as','its'}
        def words(s):
            raw = set(re.sub(r'[^a-z0-9]',' ',s.lower()).split()) - stops
            # Add stemmed versions (strip common suffixes) for fuzzy matching
            stemmed = set()
            for w in raw:
                stemmed.add(w)
                for suffix in ['ing','tion','tions','ed','s','es','er','ers','al','ment','ments']:
                    if w.endswith(suffix) and len(w) - len(suffix) >= 4:
                        stemmed.add(w[:-len(suffix)])
            return stemmed
        iw, gw = words(internal), words(gp)
        if not iw:
            return False
        if len(iw & gw) / len(iw) >= 0.20:
            return False
        # Low title overlap — use ALL internal inventors from the family as override
        family_inventors = " ".join(
            rr.get("Inventors","") for rr in rows_out
            if rr.get("Tech ID") == r.get("Tech ID") and rr.get("Inventors","")
        )
        # Check GP All Inventors field
        if _inventors_overlap(family_inventors, r.get("GP All Inventors","")):
            return False
        # Also check if any inventor last name appears in the GP title itself
        inv_lastnames = set(re.sub(r'[^a-z]','',p.strip().split()[-1])
                           for p in re.split(r'[;,]', family_inventors.lower()) if p.strip())
        if inv_lastnames & words(gp):
            return False
        return True

    for r in rows_out:
        if _is_mismatch(r):
            wrong_title = r.get("GP Title","unknown")
            r["Not Found Reason"] = f"⚠ Serial matched wrong patent: '{wrong_title[:80]}' — verify serial number in source data"
            r["Search Method"]        = "mismatch"
            r["Google Patents Link"]  = ""
            r["GP Title"]             = ""
            r["GP Status"]            = ""
            r["GP Filing Date"]       = ""
            r["GP Priority Date"]     = ""
            r["GP Issue Date"]        = ""
            r["GP Est. Exp. Date"]    = ""
            r["GP Assignee"]          = ""
            r["GP All Assignees"]     = ""
            r["GP Inventors"]         = ""
            r["GP All Inventors"]     = ""
            r["GP WO Pub. No."]       = ""
            r["Abstract (snippet)"]   = ""
            r["gp_claims"]            = ""
            r["gp_description"]       = ""

    # ── Excel ──────────────────────────────────────────────────────────────────
    res_df = pd.DataFrame(rows_out)
    wb = Workbook()

    # ── Sheet 1: All applications ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Applications"

    hdr_fill  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    fam_fills = [
        PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB"),  # light blue
        PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF"),  # white
    ]

    for ci, col in enumerate(res_df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    status_colors = {
        "active":      "C6EFCE", "granted":  "C6EFCE", "issued": "C6EFCE",
        "pending":     "FFEB9C", "filed":    "FFEB9C",
        "expired":     "FFC7CE", "abandoned":"FFC7CE",
        "provisional": "FFD9B3",
    }

    # Alternate row background per family for visual grouping
    fam_order = list(dict.fromkeys(r["Tech ID"] for r in rows_out))
    fam_idx   = {tid: i for i, tid in enumerate(fam_order)}

    DATE_COLS = {"GP Filing Date", "GP Priority Date", "GP Issue Date", "GP Est. Exp. Date"}
    DATE_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")  # soft green
    DATE_FONT_COLOR = "375623"  # dark green

    for ri, row_data in enumerate(rows_out, 2):
        fam_fill = fam_fills[fam_idx[row_data["Tech ID"]] % 2]
        for ci, col in enumerate(res_df.columns, 1):
            val = row_data.get(col,"") or ""
            is_link = col == "Google Patents Link" and str(val).startswith("http")
            c = ws.cell(row=ri, column=ci, value="View Patent" if is_link else val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="top", wrap_text=False)
            c.fill = fam_fill

            if is_link:
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                c.fill = fam_fill
                continue

            if col in ("Status (Internal)", "GP Status"):
                for kw, color in status_colors.items():
                    if kw in str(val).lower():
                        c.fill = PatternFill("solid", start_color=color, end_color=color)
                        break

            # Highlight date cells that have data
            if col in DATE_COLS and val:
                c.fill = DATE_FILL
                c.font = Font(name="Arial", size=9, color=DATE_FONT_COLOR)

            # Soft amber flag on Not Found Reason for individual members
            # (only within families that have at least one other member found)
            if col == "Not Found Reason" and val:
                fam_has_any_found = any(
                    r["Search Method"] not in ("not_found", "provisional_unpublished", "wipo_link_only")
                    for r in rows_out if r["Tech ID"] == row_data["Tech ID"]
                )
                if fam_has_any_found:
                    c.fill = PatternFill("solid", start_color="FFE699", end_color="FFE699")
                    c.font = Font(name="Arial", size=9, color="7F6000")

    widths = {
        "Tech ID":14,"Family Size":10,"Filing Type":18,
        "Title (Original)":40,"Serial Number":16,"Country":18,
        "Status (Internal)":16,"Status Date (Internal)":16,"Patent Number":14,
        "Publication Number":18,"File Date":12,"Inventors":28,"Licensing Manager":18,
        "GP Title":40,"GP Status":24,
        "GP Filing Date":14,"GP Priority Date":14,"GP Issue Date":14,"GP Est. Exp. Date":14,
        "GP Assignee":24,"GP All Assignees":28,"GP Inventors":28,"GP All Inventors":28,
        "Google Patents Link":14,"Search Method":20,
        "Not Found Reason":40,"Abstract (snippet)":50,
    }
    for ci, col in enumerate(res_df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Family Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Family Summary")
    sum_cols = [
        "Tech ID", "Family Size", "Invention (Title)", "All Titles in Family",
        "Invention Summary",
        "Inventors (Internal)", "Co-Inventors (GP)",
        "Primary Assignee", "Co-Assignees (GP)",
        "Licensing Manager", "Filings", "Has PCT?", "PCT Publication No.",
        "Overall Status",
        "Earliest Priority Date", "Earliest Filing Date", "Latest Issue Date", "Latest Est. Exp. Date",
        "Patent Numbers", "Best GP Link",
        "⚠ Data Gap?", "Reason Not Found",
    ]
    for ci, col in enumerate(sum_cols, 1):
        c = ws2.cell(row=1, column=ci, value=col)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    STATUS_COLORS = {
        "granted":   ("C6EFCE", "276221"),
        "active":    ("C6EFCE", "276221"),
        "issued":    ("C6EFCE", "276221"),
        "allowed":   ("C6EFCE", "276221"),
        "pending":   ("FFEB9C", "7D6608"),
        "filed":     ("FFEB9C", "7D6608"),
        "published": ("DDEBF7", "1F4E79"),
        "expired":   ("FCE4D6", "843C0C"),
        "abandoned": ("FFC7CE", "9C0006"),
        "withdrawn": ("FFC7CE", "9C0006"),
        "provisional":("FFD9B3","7F3F00"),
    }
    FLAG_FILL   = PatternFill("solid", start_color="FF0000", end_color="FF0000")
    FLAG_FONT   = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    ORANGE_FILL = PatternFill("solid", start_color="F4B942", end_color="F4B942")
    ORANGE_FONT = Font(name="Arial", size=9, bold=True, color="FFFFFF")

    def status_fill_font(status_str):
        s = str(status_str).lower()
        for kw, (bg, fg) in STATUS_COLORS.items():
            if kw in s:
                return (PatternFill("solid", start_color=bg, end_color=bg),
                        Font(name="Arial", size=9, bold=True, color=fg))
        return None, None

    ri2 = 2
    family_summaries = {}  # tech_id -> {title, patent_nums, summary}
    for tech_id in fam_order:
        fam_rows = [r for r in rows_out if r["Tech ID"] == tech_id]

        # Best title: longest non-provisional
        titles = [r["Title (Original)"] for r in fam_rows
                  if r["Filing Type"] != "Provisional (US)"
                  and r["Title (Original)"] not in ("", "nan")]
        best_title = max(titles, key=len) if titles else (fam_rows[0]["Title (Original)"] if fam_rows else "")

        # All distinct titles in the family (numbered if more than one)
        all_titles_list = list(dict.fromkeys(
            r["Title (Original)"] for r in fam_rows
            if r["Title (Original)"] not in ("", "nan")
        ))
        all_titles = (" | ").join(f"{i+1}. {t}" for i, t in enumerate(all_titles_list)) \
                     if len(all_titles_list) > 1 else (all_titles_list[0] if all_titles_list else "")

        # Internal inventors — normalise to a set of last names for dedup
        inventors_internal = fam_rows[0]["Inventors"] or ""
        internal_names_lower = set(
            n.strip().lower()
            for n in inventors_internal.split(",") if n.strip()
        )

        # Co-inventors from GP — exclude anyone already in the internal list
        all_gp_inventors = []
        for r in fam_rows:
            for name in (r.get("GP All Inventors", "") or "").split(","):
                n = name.strip()
                if n and n.lower() not in internal_names_lower and n not in all_gp_inventors:
                    all_gp_inventors.append(n)
        co_inventors = ", ".join(all_gp_inventors) if all_gp_inventors else "—"

        # Assignees — first is primary, rest go to co-assignees
        all_gp_assignees = []
        for r in fam_rows:
            for name in (r.get("GP All Assignees", "") or "").split(","):
                n = name.strip()
                if n and n not in all_gp_assignees:
                    all_gp_assignees.append(n)
        primary_assignee = all_gp_assignees[0] if all_gp_assignees else "—"
        co_assignees     = ", ".join(all_gp_assignees[1:]) if len(all_gp_assignees) > 1 else "—"

        lic_mgr = fam_rows[0]["Licensing Manager"] or ""
        filings = ", ".join(r["Filing Type"] for r in fam_rows)
        has_pct = "Yes" if (
            any(r["Filing Type"] == "PCT (International)" for r in fam_rows) or
            any(r.get("GP WO Pub. No.", "") for r in fam_rows) or
            any(str(r.get("Publication Number","")).upper().startswith("WO") for r in fam_rows)
        ) else "No"

        # Overall family status — most advanced non-provisional
        status_priority = ["granted","active","issued","allowed","pending","filed",
                           "published","expired","abandoned","withdrawn","provisional"]
        family_statuses = [
            (r["GP Status"] or r["Status (Internal)"] or "").lower()
            for r in fam_rows if r["Filing Type"] != "Provisional (US)"
        ]
        overall_status = "Unknown"
        for sp in status_priority:
            if any(sp in s for s in family_statuses):
                overall_status = sp.capitalize()
                break

        pat_nums  = ", ".join(
            str(r["Patent Number"]) for r in fam_rows
            if r["Patent Number"] and str(r["Patent Number"]) not in ("nan", "")
        )
        best_link = next(
            (r["Google Patents Link"] for r in fam_rows
             if r["Google Patents Link"] and r["Filing Type"] != "Provisional (US)"), ""
        )

        # Aggregate dates across family — earliest priority/filing, latest issue/expiration
        def best_date(rows, field, mode="min"):
            """Return min or max non-empty date string across rows."""
            vals = [r.get(field,"") for r in rows if r.get(field,"") and r.get(field,"") != "nan"]
            if not vals:
                return ""
            try:
                return (min if mode == "min" else max)(vals)
            except Exception:
                return vals[0]

        earliest_priority = _d(best_date(fam_rows, "GP Priority Date", "min"))
        earliest_filing   = _d(best_date(fam_rows, "GP Filing Date",   "min"))
        latest_issue      = _d(best_date(fam_rows, "GP Issue Date",    "max"))
        latest_expiration = _d(best_date(fam_rows, "GP Est. Exp. Date","max"))

        # Data gap logic — three distinct cases:
        # 1. Family is ALL provisionals → flag as "provisional only, no public filing yet"
        # 2. Family has non-provisionals but some couldn't be found → data gap
        # 3. Family has a granted patent found → no gap (grant covers the family)
        all_provisional = all(r["Filing Type"] == "Provisional (US)" for r in fam_rows)
        family_has_grant = any(
            "granted" in (r["GP Status"] or r["Status (Internal)"]).lower()
            or "issued" in (r["Status (Internal)"]).lower()
            for r in fam_rows
        )
        # Any non-provisional member that was successfully found on GP
        # (mismatched rows already cleared above before sheet writing)
        mismatched_rows = []  # already handled in pre-processing pass

        family_has_any_found = any(
            r.get("Google Patents Link", "")
            for r in fam_rows
            if r["Filing Type"] != "Provisional (US)"
            and r not in mismatched_rows
        )
        gap_rows = [
            r for r in fam_rows
            if r["Filing Type"] != "Provisional (US)"
            and (
                not r.get("Google Patents Link", "")
                and r["Search Method"] in ("not_found", "wipo_link_only", "mismatch")
            )
        ]

        if all_provisional:
            overall_status = "Provisional only"
            has_gap  = True
            gap_flag = "⚠ YES"
            gap_reasons = "Family consists entirely of provisional application(s) — no public filing exists yet. A non-provisional or PCT must be filed to enter public record."
        elif len(gap_rows) > 0 and not family_has_any_found:
            has_gap  = True
            gap_flag = "⚠ YES"
            # Separate mismatch notes from regular not-found notes
            all_reasons = list(dict.fromkeys(
                r.get("Not Found Reason", "") for r in gap_rows if r.get("Not Found Reason", "")
            ))
            gap_reasons = "; ".join(all_reasons)
        elif len(gap_rows) > 0 and family_has_any_found:
            # Some members found, some not — partial gap, softer flag
            has_gap  = False
            gap_flag = "⚠ partial"
            found_titles = [
                r.get("GP Title") or r.get("Title (Original)", "Unknown")
                for r in fam_rows
                if r.get("Google Patents Link", "")
                and r["Filing Type"] != "Provisional (US)"
                and r not in mismatched_rows
            ]
            not_found_titles = [
                r.get("Title (Original)", "Unknown")
                for r in gap_rows
                if r not in mismatched_rows
            ]
            mismatch_titles = [
                f"{r.get('Title (Original)','?')} → GP shows: '{r.get('GP Title','?')[:50]}'"
                for r in mismatched_rows
            ]
            found_str     = "✓ Found: " + " | ".join(found_titles) if found_titles else ""
            not_found_str = "✗ Not found: " + " | ".join(not_found_titles) if not_found_titles else ""
            mismatch_str  = "⚠ Title mismatch (check serial#): " + " | ".join(mismatch_titles) if mismatch_titles else ""
            gap_reasons = "; ".join(filter(None, [found_str, not_found_str, mismatch_str]))
        else:
            has_gap     = False
            gap_flag    = ""
            gap_reasons = ""

        # PCT publication numbers — source data first, then GP-scraped from any family member
        wo_from_source = [
            r.get("Publication Number","") for r in fam_rows
            if r["Filing Type"] == "PCT (International)"
            and r.get("Publication Number","") not in ("","nan")
        ]
        wo_from_gp = [
            r.get("GP WO Pub. No.","") for r in fam_rows
            if r.get("GP WO Pub. No.","") not in ("","nan")
        ]
        pct_pub_nums = ", ".join(dict.fromkeys(filter(None, wo_from_source + wo_from_gp)))

        # Generate NLP invention summary from scraped patent text
        inv_summary = ""
        if os.getenv("ANTHROPIC_API_KEY", ""):
            # Collect best available text from found family members
            best_abstract    = next((r.get("Abstract (snippet)","") for r in fam_rows if r.get("Abstract (snippet)","")), "")
            best_claims      = next((r.get("gp_claims","") for r in fam_rows if r.get("gp_claims","")), "")
            best_description = next((r.get("gp_description","") for r in fam_rows if r.get("gp_description","")), "")

            # If no text from GP, try EPO OPS API for PCT families (much more reliable than scraping)
            if not any([best_abstract, best_claims, best_description]):
                wo_num = next((
                    r.get("GP WO Pub. No.", "") or r.get("Publication Number", "")
                    for r in fam_rows
                    if (r.get("GP WO Pub. No.", "") or r.get("Publication Number", "")).upper().startswith("WO")
                ), "")
                if wo_num:
                    print(f"  Fetching abstract via EPO OPS for {wo_num}...")
                    epo_abstract = epo_get_abstract(wo_num)
                    if epo_abstract and len(epo_abstract) > 80:
                        best_abstract = epo_abstract
                        print(f"    EPO abstract: {len(best_abstract)} chars")

            if any([best_abstract, best_claims, best_description]):
                print(f"  Generating invention summary for {tech_id} (abstract:{len(best_abstract)} claims:{len(best_claims)} desc:{len(best_description)})...")
                inv_summary = generate_invention_summary(
                    best_title, pat_nums, best_abstract, best_claims, best_description
                )

        family_summaries[tech_id] = {
            "title":       best_title,
            "patent_nums": pat_nums,
            "summary":     inv_summary,
        }

        row_vals = [
            tech_id, len(fam_rows), best_title, all_titles,
            inv_summary,
            inventors_internal, co_inventors,
            primary_assignee, co_assignees,
            lic_mgr, filings, has_pct, pct_pub_nums,
            overall_status,
            earliest_priority, earliest_filing, latest_issue, latest_expiration,
            pat_nums, best_link,
            gap_flag, gap_reasons,
        ]

        PURPLE_FILL = PatternFill("solid", start_color="D9B3FF", end_color="D9B3FF")
        PURPLE_FONT = Font(name="Arial", size=9, bold=True, color="4B0082")

        fam_fill2 = fam_fills[fam_idx[tech_id] % 2]
        for ci, val in enumerate(row_vals, 1):
            col_name = sum_cols[ci - 1]
            is_link = col_name == "Best GP Link" and str(val).startswith("http")
            c = ws2.cell(row=ri2, column=ci, value="View Patent" if is_link else val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.fill = fam_fill2

            if is_link:
                c.hyperlink = val
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                continue

            if col_name == "Overall Status":
                if val == "Provisional only":
                    c.fill = PURPLE_FILL
                    c.font = PURPLE_FONT
                else:
                    sf, ff = status_fill_font(val)
                    if sf:
                        c.fill = sf
                        c.font = ff

            if col_name == "Has PCT?" and val == "Yes":
                c.fill = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
                c.font = Font(name="Arial", size=9, bold=True, color="1F4E79")

            if col_name == "⚠ Data Gap?" and val:
                if val == "⚠ partial":
                    c.fill = ORANGE_FILL
                    c.font = ORANGE_FONT
                else:
                    c.fill = FLAG_FILL
                    c.font = FLAG_FONT

            if col_name == "Reason Not Found" and val:
                c.fill = ORANGE_FILL
                c.font = ORANGE_FONT

        ri2 += 1

    sum_widths = [14, 10, 42, 50, 60, 30, 30, 24, 28, 18, 36, 10, 18, 16, 14, 14, 14, 14, 20, 14, 10, 55]
    for ci, w in enumerate(sum_widths, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(sum_cols))}1"

    wb.save(OUTPUT_XLSX)
    print(f"Saved {OUTPUT_XLSX}  (2 sheets: 'All Applications' + 'Family Summary')")
    print(f"\n── Total Google Patents requests: {scrape_count} ──")

    # ── Text summary ───────────────────────────────────────────────────────────
    now          = datetime.now().strftime("%Y-%m-%d %H:%M")
    provisionals = [r for r in rows_out if r["Search Method"] == "provisional_unpublished"]
    found        = [r for r in rows_out if r["Search Method"] not in ("not_found","provisional_unpublished")]
    not_found    = [r for r in rows_out if r["Search Method"] == "not_found"]
    by_method    = {}
    for r in rows_out:
        by_method[r["Search Method"]] = by_method.get(r["Search Method"],0)+1
    sc = res_df["Status (Internal)"].value_counts()
    n_families = len(fam_order)

    lines = [
        "="*70,
        "  TJU PATENT PORTFOLIO — GOOGLE PATENTS LOOKUP SUMMARY",
        f"  Generated: {now}",
        f"  Method: Direct Google Patents scraping (no API key required)",
        "="*70,"",
        "DATA QUALITY NOTE","-"*40,
        f"  Rows in source file:           {len(df)}",
        f"  Exact duplicate rows:          {n_dupes}  ← lookup ran on unique records only",
        f"  Unique patent applications:    {len(df_u)}",
        f"  Distinct patent families:      {n_families}  (grouped by Tech ID)","",
        "NOTE ON PROVISIONALS","-"*40,
        f"  {len(provisionals)} provisional application(s) (63/xxx, 62/xxx) detected.",
        f"  Provisionals are NEVER published by law and cannot appear in any",
        f"  public patent database. This is expected and not an error.",
        "",
        "LOOKUP RESULTS","-"*40,
        f"  Found on Google Patents:       {len(found)} / {len(df_u)}",
        f"  Provisionals (unpublishable):  {len(provisionals)}",
        f"  Not found:                     {len(not_found)}",
        f"  Google Patents requests made:  {scrape_count}","",
    ]
    for method, count in by_method.items():
        lines.append(f"    {method:<30} {count}")
    lines.append("")

    if not_found:
        lines += ["NOT FOUND (excluding provisionals):","-"*40]
        for r in not_found:
            lines.append(f"  • {r['Tech ID']:<18} {r['Serial Number'] or 'no serial':<20} {r['Title (Original)'][:50]}")
        lines.append("")

    lines += ["PORTFOLIO STATUS BREAKDOWN (internal data)","-"*40]
    for status, count in sc.items():
        lines.append(f"  {status:<34} {count}")

    # Family-level breakdown
    lines += ["","PATENT FAMILY BREAKDOWN","-"*40]
    lines.append(f"  Total distinct families:          {n_families}")

    # Count families by overall GP status
    fam_statuses = {}
    fam_has_pct  = 0
    fam_all_prov = 0
    fam_has_grant = 0
    fam_pending  = 0
    fam_not_found = 0
    for tech_id in fam_order:
        fam_rows = [r for r in rows_out if r["Tech ID"] == tech_id]
        all_prov = all(r["Filing Type"] == "Provisional (US)" for r in fam_rows)
        has_pct  = any(r["Filing Type"] == "PCT (International)" for r in fam_rows)
        has_grant = any(
            "granted" in (r["GP Status"] or "").lower() or
            "active"  in (r["GP Status"] or "").lower() or
            "issued"  in (r["Status (Internal)"] or "").lower()
            for r in fam_rows
        )
        has_pending = any(
            "pending" in (r["Status (Internal)"] or "").lower() or
            "filed"   in (r["Status (Internal)"] or "").lower()
            for r in fam_rows
            if r["Filing Type"] != "Provisional (US)"
        )
        any_found = any(r.get("Google Patents Link","") for r in fam_rows)

        if all_prov:       fam_all_prov  += 1
        if has_pct:        fam_has_pct   += 1
        if has_grant:      fam_has_grant += 1
        elif has_pending:  fam_pending   += 1
        if not any_found and not all_prov: fam_not_found += 1

    fam_has_design = 0
    for tech_id in fam_order:
        fam_rows = [r for r in rows_out if r["Tech ID"] == tech_id]
        if any(r["Filing Type"] == "Design (US)" for r in fam_rows):
            fam_has_design += 1

    lines.append(f"  Families with granted patent:     {fam_has_grant}")
    lines.append(f"  Families with pending filings:    {fam_pending}")
    lines.append(f"  Provisional-only families:        {fam_all_prov}")
    lines.append(f"  Families with PCT filing:         {fam_has_pct}")
    lines.append(f"  Families with design patent:      {fam_has_design}")
    lines.append(f"  Families not found on GP:         {fam_not_found}")

    lines += ["","","PATENT FAMILIES — DETAIL","="*70]
    for tech_id in fam_order:
        fam_rows = [r for r in rows_out if r["Tech ID"] == tech_id]
        titles = [r["Title (Original)"] for r in fam_rows
                  if r["Filing Type"] != "Provisional (US)" and r["Title (Original)"] not in ("","nan")]
        best_title = max(titles, key=len) if titles else fam_rows[0]["Title (Original)"]
        inventors  = fam_rows[0]["Inventors"] or "—"
        lic_mgr    = fam_rows[0]["Licensing Manager"] or "—"

        lines += [
            "",
            f"  {tech_id}  ({len(fam_rows)} filing{'s' if len(fam_rows)>1 else ''})",
            f"  Invention:  {best_title[:65]}",
            f"  Inventors:  {inventors}",
            f"  Manager:    {lic_mgr}",
            f"  {'Filing Type':<22} {'Serial':<22} {'Status':<28} {'GP Link'}",
            f"  {'-'*22} {'-'*22} {'-'*28} {'-'*30}",
        ]
        for r in fam_rows:
            status = r["GP Status"] or r["Status (Internal)"] or "—"
            link   = r["Google Patents Link"] or "—"
            lines.append(
                f"  {r['Filing Type']:<22} {(r['Serial Number'] or '—'):<22} {status:<28} {link}"
            )

    lines += ["","="*70, f"  Excel output: {OUTPUT_XLSX} (sheets: All Applications | Family Summary)",
              "="*70]

    # Invention summaries section — only if any were generated
    summaries_with_content = {k: v for k, v in family_summaries.items() if v.get("summary")}
    if summaries_with_content:
        lines += ["","","INVENTION SUMMARIES","="*70,
                  "  AI-generated summaries from patent abstracts, claims, and descriptions.",
                  "  Technical summary followed by plain English summary.",
                  ""]
        for tech_id, info in summaries_with_content.items():
            lines += [
                f"  {tech_id}",
                f"  Title:   {info['title'][:75]}",
                f"  Patent:  {info['patent_nums'] or 'Pending'}",
                "",
            ]
            # Indent each line of the summary
            for line in info["summary"].splitlines():
                lines.append(f"    {line}")
            lines.append("")
        lines.append("="*70)

    with open(OUTPUT_TXT,"w",encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"Saved {OUTPUT_TXT}")
    print(f"\nDone! Runtime estimate: ~{len(df_u)*5//60} min for {len(df_u)} records at ~5s/record.")


if __name__ == "__main__":
    main()
